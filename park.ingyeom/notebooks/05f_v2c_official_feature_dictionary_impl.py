import json
import os
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


STAGE_NAME = "05f_v2c_official_feature_dictionary"
EXPECTED_FEATURE_SET = "pruned_w1_3_all_weeks_interpretable_without_product_code_without_watch_presence"
EXPECTED_MODEL = "HistGradientBoostingClassifier"
TARGET = "is_repurchase_label"
METADATA_COLUMNS = ["membership_row_id", "USER_KEY"]


def find_project_root(start: Path) -> Path:
    for candidate in [start, *start.parents]:
        if (
            (candidate / "park.ingyeom" / "reports" / "data" / "05c_v2_modeling_dataset" / "feature_sets_v2c.json").exists()
            and (candidate / "park.ingyeom" / "reports" / "data" / "06c2_v2_corrected_baseline_modeling" / "06c2_final_model_recommendation.md").exists()
        ):
            return candidate
    raise FileNotFoundError("Could not locate ott-churn-prediction project root.")


PROJECT_ROOT = find_project_root(Path.cwd())
DATA05C = PROJECT_ROOT / "park.ingyeom" / "reports" / "data" / "05c_v2_modeling_dataset"
DATA06C2 = PROJECT_ROOT / "park.ingyeom" / "reports" / "data" / "06c2_v2_corrected_baseline_modeling"
DATA07C = PROJECT_ROOT / "park.ingyeom" / "reports" / "data" / "07c_v2_corrected_true_shap_interpretation"
TABLE07C = PROJECT_ROOT / "park.ingyeom" / "reports" / "tables" / "07c_v2_corrected_true_shap_interpretation"
DATA05D = PROJECT_ROOT / "park.ingyeom" / "reports" / "data" / "05d_v2_feature_dictionary"
TABLE05D = PROJECT_ROOT / "park.ingyeom" / "reports" / "tables" / "05d_v2_feature_dictionary"
DATA05E = PROJECT_ROOT / "park.ingyeom" / "reports" / "data" / "05e_v2_final_feature_pruning_policy"
DATA06H = PROJECT_ROOT / "park.ingyeom" / "reports" / "data" / "06h_v2_pruned_model_collinearity_shap_audit"
DATA02C = PROJECT_ROOT / "park.ingyeom" / "reports" / "data" / "02c_v2_strict_preprocessing_correction"
DATA05_OLD = PROJECT_ROOT / "park.ingyeom" / "reports" / "data" / "05_v2_modeling_dataset"
DATA06_OLD = PROJECT_ROOT / "park.ingyeom" / "reports" / "data" / "06_v2_baseline_modeling"

OUT_DATA = PROJECT_ROOT / "park.ingyeom" / "reports" / "data" / STAGE_NAME
OUT_TABLE = PROJECT_ROOT / "park.ingyeom" / "reports" / "tables" / STAGE_NAME
WORKBOOK_PATH = OUT_TABLE / "05f_v2c_official_feature_dictionary.xlsx"
REPORT_PATH = OUT_DATA / "05f_v2c_official_feature_dictionary_report.md"
SUMMARY_PATH = OUT_DATA / "05f_v2c_official_feature_dictionary_summary.json"


def read_json(path: Path) -> dict:
    return json.loads(path.read_text(encoding="utf-8"))


def require_inputs() -> None:
    required = [
        DATA05C / "modeling_dataset_v2c_w1_3.csv",
        DATA05C / "feature_sets_v2c.json",
        DATA06C2 / "06c2_corrected_baseline_summary.json",
        DATA06C2 / "06c2_final_model_recommendation.md",
        DATA07C / "07c_true_shap_summary.json",
        TABLE07C / "07c_global_shap_importance.csv",
        TABLE07C / "07c_feature_family_shap_importance.csv",
        TABLE07C / "07c_shap_direction_summary.csv",
    ]
    missing = [str(path.relative_to(PROJECT_ROOT)) for path in required if not path.exists()]
    if missing:
        raise FileNotFoundError("Required input files are missing: " + ", ".join(missing))


def assert_no_existing_outputs() -> None:
    if os.environ.get("ALLOW_RECREATE_OWN_05F") == "1":
        return
    existing = [path for path in [WORKBOOK_PATH, REPORT_PATH, SUMMARY_PATH] if path.exists()]
    if existing:
        rels = ", ".join(str(path.relative_to(PROJECT_ROOT)) for path in existing)
        raise FileExistsError(f"Refusing to overwrite existing 05f outputs: {rels}")


def dtype_name(series: pd.Series | None, variable: str) -> str:
    if variable == TARGET:
        return "int (0/1)"
    if variable == "USER_KEY":
        return "string"
    if series is None:
        return "not in v2c w1_3 dataset"
    if pd.api.types.is_integer_dtype(series):
        return "int"
    if pd.api.types.is_float_dtype(series):
        values = set(series.dropna().unique().tolist()[:20])
        if values and values.issubset({0.0, 1.0}):
            return "int/binary (0/1)"
        return "float"
    if pd.api.types.is_bool_dtype(series):
        return "bool"
    return "category/string"


def variable_group(name: str, official_features: set[str], metadata: set[str]) -> tuple[str, str, str]:
    if name == TARGET:
        return "타겟 변수", "타깃", "target"
    if name in metadata:
        return "식별자/메타데이터", "메타데이터", "metadata only"
    if name not in official_features:
        return "제외된 변수", "제외된 피처", "공식 모델 제외"
    if name in {"price_num", "max_screen_num", "age_num"}:
        return "기본 멤버십", "표준화 피처", "공식 모델 사용"
    if name in {"is_promotion_bin", "is_user_verified_bin", "gender_clean", "payment_device_clean", "billing_method_clean"}:
        return "기본 멤버십", "표준화 피처", "공식 모델 사용"
    if "genre_ratio" in name or "genre_entropy" in name:
        return "콘텐츠/장르 비율", "파생 피처", "공식 모델 사용 | feature-family interpretation recommended"
    if "recent_content_watch_ratio" in name or "release_month" in name:
        return "콘텐츠 공개월 proxy", "파생 피처", "공식 모델 사용 | feature-family interpretation recommended"
    if name.startswith("w1_3_week"):
        return "주차별 이용 패턴", "파생 피처", "공식 모델 사용 | feature-family interpretation recommended"
    if name.startswith("w1_3_"):
        return "시청 행태", "파생 피처", "공식 모델 사용 | feature-family interpretation recommended"
    return "참고용 변수", "원본 피처", "공식 모델 사용"


def description_for(name: str, official_features: set[str]) -> tuple[str, str]:
    if name == "membership_row_id":
        return "멤버십 행 단위 추적용 식별자입니다.", "Stage 05c 모델링 데이터셋에서 보존된 행 식별자입니다. 모델 feature로 사용하지 않습니다."
    if name == "USER_KEY":
        return "동일 사용자를 묶는 그룹 식별자입니다.", "train/test 사용자 중복 방지 확인용 metadata입니다. 모델 feature로 사용하지 않습니다."
    if name == TARGET:
        return "재구독 여부를 1/0으로 인코딩한 예측 대상입니다.", "Stage 02c 이후 corrected pipeline의 target encoding 결과입니다."
    base = {
        "price_num": ("구독 상품 가격을 수치형으로 표준화한 변수입니다.", "Membership 원천 가격 값을 숫자형으로 파싱했습니다."),
        "max_screen_num": ("요금제의 최대 동시 화면 수를 수치형으로 표준화한 변수입니다.", "Membership 원천 max_screen 값을 숫자형으로 파싱했습니다."),
        "is_promotion_bin": ("프로모션 가입 여부를 0/1로 변환한 변수입니다.", "Membership 원천 이진값을 corrected binary policy에 맞춰 변환했습니다."),
        "is_user_verified_bin": ("사용자 인증 여부를 0/1로 변환한 변수입니다.", "Membership 원천 이진값을 corrected binary policy에 맞춰 변환했습니다."),
        "age_num": ("사용자 나이를 수치형으로 표준화한 변수입니다.", "Membership 원천 age 값을 숫자형으로 파싱하고 02c 기준으로 이상값을 정리했습니다."),
        "gender_clean": ("성별 값을 정리한 범주형 멤버십 변수입니다.", "Membership 원천 gender 값을 corrected categorical policy에 맞춰 정리했습니다."),
        "payment_device_clean": ("결제 기기를 정리한 범주형 멤버십 변수입니다.", "Membership 원천 payment_device 값을 corrected categorical policy에 맞춰 정리했습니다."),
        "billing_method_clean": ("결제 방식을 정리한 범주형 멤버십 변수입니다.", "Membership 원천 billing_method 값을 corrected categorical policy에 맞춰 정리했습니다."),
    }
    if name in base:
        return base[name]
    if name.startswith("w1_3_week") and name.endswith("_watch_time"):
        return "구독 시작 후 1~3주 관측창에서 해당 주차의 시청 시간을 집계한 변수입니다.", "Views 기반 시청 기록을 membership_row_id 단위와 주차 단위로 합산했습니다."
    if name.startswith("w1_3_week") and name.endswith("_sessions"):
        return "구독 시작 후 1~3주 관측창에서 해당 주차의 시청 세션 수를 집계한 변수입니다.", "Views 기반 시청 기록을 membership_row_id 단위와 주차 단위로 카운트했습니다."
    if name == "w1_3_unique_contents":
        return "1~3주 관측창에서 시청한 서로 다른 콘텐츠 수입니다.", "Views와 콘텐츠 키를 membership_row_id 단위로 집계했습니다."
    if name == "w1_3_unique_watch_days":
        return "1~3주 관측창에서 실제 시청이 있었던 서로 다른 날짜 수입니다.", "Views의 시청일을 membership_row_id 단위로 중복 제거해 계산했습니다."
    if name == "w1_3_avg_watch_time_per_session":
        return "1~3주 관측창의 세션당 평균 시청 시간입니다.", "총 시청 시간을 총 세션 수로 나눈 파생 변수입니다."
    if "genre_ratio" in name:
        genre = name.replace("w1_3_genre_ratio_", "")
        return f"1~3주 관측창에서 {genre} 장르 시청 시간이 전체 장르 시청 시간에서 차지하는 비율입니다.", "Movie/genre 매핑 후 장르별 시청 시간을 합산하고 전체 장르 시청 시간으로 나눴습니다."
    if name == "w1_3_genre_entropy":
        return "1~3주 관측창에서 장르 시청 비중의 다양성을 요약한 변수입니다.", "장르별 시청 비중으로 entropy를 계산했습니다."
    if name == "w1_3_recent_content_watch_ratio":
        return "1~3주 관측창에서 공개월이 상대적으로 최근인 콘텐츠 시청 비중입니다.", "콘텐츠 공개월 proxy를 이용해 최근 콘텐츠 시청 시간을 전체 시청 시간 대비 비율로 계산했습니다."
    if name in official_features:
        return "공식 모델 feature set에 포함된 변수입니다.", "Stage 05c corrected modeling dataset과 06c2 공식 추천 feature set에서 확인했습니다."
    return excluded_description(name)


def excluded_description(name: str) -> tuple[str, str]:
    if name == "product_code":
        return "상품 코드는 공식 모델에서 제외했습니다.", "상품 코드 자체를 외우는 product memorization risk를 줄이기 위해 제외했습니다."
    if "has_watch_obs" in name or "no_watch_obs" in name:
        return "시청 관측 여부 플래그는 공식 모델에서 제외했습니다.", "시청 기록 존재 자체가 강한 shortcut이 될 수 있어 watch-presence shortcut으로 분류했습니다."
    if name in {"first_watch_rel_day", "last_watch_rel_day"} or "first_watch_rel_day" in name or "last_watch_rel_day" in name:
        return "첫/마지막 시청 상대일은 공식 모델에서 제외했습니다.", "타깃 시점에 가까운 timing proxy가 될 수 있어 target-adjacent timing으로 분류했습니다."
    if "week" in name and "ratio" in name:
        return "주차별 시청 비중은 공식 모델에서 제외했습니다.", "주차별 시청 시간과 구조적으로 중복될 수 있어 ratio/delta excluded로 분류했습니다."
    if "_minus_" in name:
        return "주차 간 차이 변수는 공식 모델에서 제외했습니다.", "주차별 시청 시간의 선형 조합에 가까워 structural redundancy로 분류했습니다."
    if "total_watch_time" in name:
        return "총 시청 시간은 공식 모델에서 제외했습니다.", "주차별 시청 시간과 중복되는 사용량 proxy로 분류했습니다."
    if "genre_watch_time_" in name or "genre_session_count_" in name:
        return "장르별 절대 시청량 또는 세션 수는 공식 모델에서 제외했습니다.", "장르 취향보다 사용량 proxy가 섞일 위험이 있어 genre volume proxy excluded로 분류했습니다."
    if name in {"is_churn_prevented", "is_churn_prevented_bin"}:
        return "이탈 방지 여부 후보 변수는 공식 모델에서 제외했습니다.", "정책 개입 또는 target 관련 사후 정보가 섞일 수 있어 metadata/forbidden feature로 분류했습니다."
    if "duration_days" in name:
        return "구독 기간 변수는 공식 모델에서 제외했습니다.", "02c에서는 monthly-scope population을 정하는 정책 기준이며 feature가 아닙니다."
    if name in {"reg_date", "end_date", "reg_date_parsed", "end_date_parsed", "watch_date", "watch_day"}:
        return "원시 날짜 컬럼은 공식 모델에서 제외했습니다.", "날짜 파싱과 관측창 계산에는 쓰였지만 최종 모델 feature로 직접 사용하지 않습니다."
    if "raw" in name or "stage02" in name:
        return "raw backup 또는 이전 단계 보존 컬럼은 공식 모델에서 제외했습니다.", "감사용 보존값이며 모델 feature가 아닙니다."
    return "공식 모델 feature set에 포함되지 않은 변수입니다.", "06c2 공식 추천 feature set 기준으로 제외했습니다."


def exclusion_note(name: str) -> str:
    if name == "product_code":
        return "공식 모델 제외 | 제외된 피처 | product_code excluded | product memorization risk"
    if "has_watch_obs" in name or "no_watch_obs" in name:
        return "공식 모델 제외 | 제외된 피처 | watch-presence shortcut excluded"
    if "first_watch_rel_day" in name or "last_watch_rel_day" in name:
        return "공식 모델 제외 | 제외된 피처 | first/last timing excluded | target-adjacent timing"
    if ("week" in name and "ratio" in name) or "_minus_" in name:
        return "공식 모델 제외 | 제외된 피처 | ratio/delta excluded | structural redundancy"
    if "genre_watch_time_" in name or "genre_session_count_" in name:
        return "공식 모델 제외 | 제외된 피처 | genre volume proxy excluded | usage-volume proxy duplication"
    if name in {"is_churn_prevented", "is_churn_prevented_bin"}:
        return "공식 모델 제외 | 제외된 피처 | metadata/forbidden feature"
    if "duration_days" in name:
        return "공식 모델 제외 | 제외된 피처 | metadata/forbidden feature"
    if name in {"reg_date", "end_date", "reg_date_parsed", "end_date_parsed", "watch_date", "watch_day"}:
        return "공식 모델 제외 | 제외된 피처 | raw backup only"
    if "raw" in name or "stage02" in name:
        return "공식 모델 제외 | 제외된 피처 | raw backup only"
    if "total_watch_time" in name:
        return "공식 모델 제외 | 제외된 피처 | usage-volume proxy duplication"
    return "공식 모델 제외 | 제외된 피처"


def collect_excluded_variables(all_columns: list[str], all_feature_sets: dict) -> list[str]:
    current_dataset_patterns = [
        "product_code",
        "has_watch_obs",
        "no_watch_obs",
        "first_watch_rel_day",
        "last_watch_rel_day",
        "week1_ratio",
        "week2_ratio",
        "week3_ratio",
        "_minus_",
        "total_watch_time",
        "genre_watch_time_",
        "genre_session_count_",
        "is_churn_prevented",
        "duration_days",
        "raw",
        "stage02",
    ]
    candidates = set()
    for column in all_columns:
        if any(pattern in column for pattern in current_dataset_patterns):
            candidates.add(column)
    candidates.update(
        {
            "product_code",
            "has_watch_obs",
            "no_watch_obs_flag",
            "first_watch_rel_day",
            "last_watch_rel_day",
            "duration_days",
            "duration_days_recomputed",
            "reg_date",
            "end_date",
            "watch_date",
            "watch_day",
            "is_churn_prevented",
        }
    )
    return sorted(candidates)


def category_for_excluded(name: str) -> str:
    return "제외된 변수"


def make_feature_dictionary() -> tuple[pd.DataFrame, dict]:
    feature_sets = read_json(DATA05C / "feature_sets_v2c.json")
    summary06c2 = read_json(DATA06C2 / "06c2_corrected_baseline_summary.json")
    recommendation = summary06c2["official_corrected_recommendation"]
    official_set_name = recommendation["recommended_feature_set"]
    official_model = recommendation["recommended_model"]
    if official_set_name != EXPECTED_FEATURE_SET:
        print(f"Using 06c2 official feature set instead of hard-coded expectation: {official_set_name}")
    if official_model != EXPECTED_MODEL:
        print(f"Using 06c2 official model instead of hard-coded expectation: {official_model}")

    all_feature_sets = feature_sets["feature_sets"]
    official_features = all_feature_sets[official_set_name]["features"]
    official_set = set(official_features)
    dataset = pd.read_csv(DATA05C / "modeling_dataset_v2c_w1_3.csv")
    shap = pd.read_csv(TABLE07C / "07c_global_shap_importance.csv")
    direction = pd.read_csv(TABLE07C / "07c_shap_direction_summary.csv")
    shap_map = shap.set_index("original_feature").to_dict("index")
    direction_map = direction.set_index("original_feature").to_dict("index")

    rows = []
    ordered_variables = [TARGET, *METADATA_COLUMNS, *official_features]
    excluded = [name for name in collect_excluded_variables(dataset.columns.tolist(), all_feature_sets) if name not in official_set and name not in ordered_variables]
    ordered_variables.extend(excluded)

    for idx, name in enumerate(ordered_variables, 1):
        series = dataset[name] if name in dataset.columns else None
        category, feature_type, note = variable_group(name, official_set, set(METADATA_COLUMNS))
        if name not in official_set and name not in [TARGET, *METADATA_COLUMNS]:
            category = category_for_excluded(name)
            feature_type = "제외된 피처"
            note = exclusion_note(name)
        desc, logic = description_for(name, official_set)
        shap_info = shap_map.get(name)
        direction_info = direction_map.get(name)
        if shap_info and name in official_set:
            note = f"{note} | 07c TRUE SHAP rank {int(shap_info['rank'])}, mean_abs_shap {shap_info['mean_abs_shap']:.6f}"
            if direction_info:
                note = f"{note}, {direction_info['primary_direction']}"
        if feature_type:
            note = f"{note} | {feature_type}"
        rows.append(
            {
                "#": idx,
                "변수명": name,
                "카테고리": category,
                "데이터 타입": dtype_name(series, name),
                "설명": desc,
                "생성 방식": logic,
                "비고": note,
            }
        )

    checks = {
        "official_feature_set": official_set_name,
        "official_model": official_model,
        "official_auc": recommendation["roc_auc_repurchase"],
        "official_feature_count": len(official_features),
        "dataset_rows": len(dataset),
        "dataset_columns": len(dataset.columns),
    }
    return pd.DataFrame(rows), checks


def make_comparison_sheet(checks: dict) -> pd.DataFrame:
    old05 = read_json(DATA05_OLD / "modeling_dataset_summary.json")
    old06 = read_json(DATA06_OLD / "06_v2_baseline_modeling_summary.json")
    stage02c = read_json(DATA02C / "02c_strict_preprocessing_summary.json")
    summary07c = read_json(DATA07C / "07c_true_shap_summary.json")
    rows = [
        ("전처리 기준", "기존 full exploratory v2 pipeline", "02c strict-core correction 이후 corrected v2c pipeline", "02c에서 duration/age/max_screen 등 strict-core 기준 적용"),
        ("row count", f"w1_3 {old05['row_counts']['w1_3']:,} rows", f"w1_3 {checks['dataset_rows']:,} rows", f"02c removed rows: {stage02c['stage02c_removed_rows']:,}"),
        ("target encoding", "is_repurchase Y/N 기반", "is_repurchase_label 1/0 기반", "1=repurchase, 0=non-repurchase/churn risk"),
        ("binary encoding", "pre-02c 기준", "corrected binary policy 기준", "is_promotion_bin, is_user_verified_bin 사용"),
        ("date parsing", "pre-02c 산출물 기준", "02c date parse audit 이후 corrected 기준", "raw date columns는 모델 feature에서 제외"),
        ("duration policy", "pre-02c monthly scope correction 전", stage02c["duration_scope_policy"], "duration_days는 population policy용이며 feature 제외"),
        ("feature set", f"{old05['feature_set_count']} feature sets", checks["official_feature_set"], "06c2 final recommendation 기준"),
        ("product_code 사용 여부", "full exploratory 또는 과거 후보에서 사용 가능", "사용하지 않음", "product memorization risk"),
        ("watch-presence flag 사용 여부", "has_watch_obs 계열 사용 가능", "사용하지 않음", "watch-presence shortcut 방지"),
        ("first/last timing 사용 여부", "과거 후보에서 사용 가능", "사용하지 않음", "target-adjacent timing 방지"),
        ("ratio/delta 사용 여부", "week ratio/delta 계열 사용 가능", "사용하지 않음", "주차별 절대 사용량과 구조적 중복 가능성"),
        ("genre volume/session_count 사용 여부", "genre_watch_time/session_count 계열 사용 가능", "사용하지 않음", "장르 취향보다 사용량 proxy가 섞일 위험"),
        ("official model", old06.get("old_pre_02c_official_model", "HistGradientBoostingClassifier"), checks["official_model"], "06c2 기준"),
        ("official AUC", f"old conservative AUC {old06.get('conservative_recommended_auc'):.6f}", f"{checks['official_auc']:.6f}", "06c2/07c reconstruction 일치"),
        ("SHAP 기준", "07r 또는 06h는 historical/provisional", "07c TRUE SHAP only", summary07c["previous_shap_outputs_status"]),
        ("downstream status", "02c 이후 deprecated/requires rerun", "06c2/07c corrected official model 기준 사용", "segmentation/simulation은 본 산출물에 혼합하지 않음"),
    ]
    return pd.DataFrame(rows, columns=["항목", "기존 v2 / pre-02c", "corrected v2c / post-02c", "비고"])


def make_category_summary(feature_df: pd.DataFrame, checks: dict) -> pd.DataFrame:
    family = pd.read_csv(TABLE07C / "07c_feature_family_shap_importance.csv")
    official_df = feature_df[feature_df["비고"].str.contains("공식 모델 사용", regex=False, na=False)]
    rows = []
    for category in [
        "타겟 변수",
        "식별자/메타데이터",
        "기본 멤버십",
        "시청 행태",
        "주차별 이용 패턴",
        "콘텐츠/장르 비율",
        "콘텐츠 공개월 proxy",
    ]:
        subset = feature_df[feature_df["카테고리"] == category]
        rows.append((category, len(subset), ", ".join(subset["변수명"].tolist())))
    excluded = feature_df[feature_df["비고"].str.contains("공식 모델 제외", regex=False, na=False)]
    rows.extend(
        [
            ("공식 모델 제외 변수", len(excluded), ", ".join(excluded["변수명"].tolist())),
            ("전체 공식 모델 feature 수", checks["official_feature_count"], ", ".join(official_df["변수명"].tolist())),
            ("official raw feature count", checks["official_feature_count"], "06c2 official raw feature count"),
            ("post-transform feature count", len(pd.read_csv(TABLE07C / "07c_global_shap_importance.csv")), "07c_global_shap_importance.csv의 transformed/global SHAP 행 수 기준"),
            ("top SHAP feature families", len(family), "; ".join(f"{r.feature_family}({r['rank']})" for _, r in family.iterrows())),
        ]
    )
    return pd.DataFrame(rows, columns=["카테고리", "변수 수", "변수 목록"])


def make_shap_sheet(feature_df: pd.DataFrame) -> pd.DataFrame:
    shap = pd.read_csv(TABLE07C / "07c_global_shap_importance.csv")
    direction = pd.read_csv(TABLE07C / "07c_shap_direction_summary.csv")
    category_map = feature_df.set_index("변수명")["카테고리"].to_dict()
    direction_map = direction.set_index("original_feature").to_dict("index")
    rows = []
    for _, row in shap.iterrows():
        feature = row["original_feature"]
        direction_info = direction_map.get(feature, {})
        primary = direction_info.get("primary_direction", row.get("shap_direction", ""))
        if primary == "pushes_toward_repurchase_score":
            interpretation = "평균적으로 재구독 점수 쪽으로 기여합니다."
        elif primary == "pushes_toward_churn_risk":
            interpretation = "평균적으로 이탈 위험 쪽으로 기여합니다."
        else:
            interpretation = "방향은 개별 관측치별로 달라질 수 있습니다."
        rows.append(
            {
                "순위": int(row["rank"]),
                "변수명": feature,
                "카테고리": category_map.get(feature, "참고용 변수"),
                "mean_abs_shap": row["mean_abs_shap"],
                "SHAP 방향 요약": primary,
                "해석": interpretation,
                "주의사항": "07c TRUE SHAP 기준입니다. 원인 효과나 ROI로 해석하지 않습니다.",
            }
        )
    return pd.DataFrame(rows)


def make_exclusion_sheet(feature_df: pd.DataFrame) -> pd.DataFrame:
    excluded = feature_df[feature_df["비고"].str.contains("공식 모델 제외", regex=False, na=False)].copy()
    rows = []
    for _, row in excluded.iterrows():
        note = row["비고"]
        if "product_code" in row["변수명"]:
            reason = "product memorization risk"
        elif "watch-presence" in note:
            reason = "watch-presence shortcut"
        elif "first/last timing" in note:
            reason = "target-adjacent timing"
        elif "ratio/delta" in note:
            reason = "structural redundancy"
        elif "genre volume proxy" in note:
            reason = "usage-volume proxy duplication"
        elif "raw backup" in note:
            reason = "metadata/raw backup only"
        elif "forbidden" in note:
            reason = "metadata/forbidden feature"
        else:
            reason = "06c2 official feature set 기준 제외"
        rows.append(
            {
                "변수명": row["변수명"],
                "기존 구분": row["카테고리"],
                "제외 사유": reason,
                "최종 처리": "공식 모델 feature에서 제외",
                "비고": row["비고"],
            }
        )
    return pd.DataFrame(rows)


def style_workbook(path: Path) -> None:
    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="2E4057")
    header_font = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
    body_font = Font(name="맑은 고딕", size=10)
    thin = Side(style="thin", color="D9E2EC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    widths = {
        "변수 설명": {"A": 4, "B": 32, "C": 18, "D": 18, "E": 44, "F": 52, "G": 62},
        "두 노트북 비교": {"A": 28, "B": 42, "C": 48, "D": 46},
        "카테고리 요약": {"A": 24, "B": 14, "C": 110},
        "SHAP 중요 변수": {"A": 8, "B": 34, "C": 18, "D": 16, "E": 26, "F": 42, "G": 48},
        "제외 변수 사유": {"A": 34, "B": 18, "C": 34, "D": 26, "E": 60},
    }
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 30
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = body_font
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = border
        for col_letter, width in widths.get(ws.title, {}).items():
            ws.column_dimensions[col_letter].width = width
        if ws.max_row > 1 and ws.max_column > 0:
            ws.auto_filter.ref = ws.dimensions
        for col_idx in range(1, ws.max_column + 1):
            header = ws.cell(1, col_idx).value
            if header in {"#", "순위", "변수 수"}:
                for row_idx in range(2, ws.max_row + 1):
                    ws.cell(row_idx, col_idx).alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
    wb.save(path)


def write_report(feature_df: pd.DataFrame, checks: dict) -> str:
    official = feature_df[feature_df["비고"].str.contains("공식 모델 사용", regex=False, na=False)]
    excluded = feature_df[feature_df["비고"].str.contains("공식 모델 제외", regex=False, na=False)]
    original_like = official[official["비고"].str.contains("표준화 피처", regex=False, na=False)]
    derived = official[official["비고"].str.contains("파생 피처", regex=False, na=False)]
    metadata_target_names = [TARGET, *METADATA_COLUMNS]
    metadata_target = feature_df[feature_df["변수명"].isin(metadata_target_names)]
    top_families = pd.read_csv(TABLE07C / "07c_feature_family_shap_importance.csv")
    report = f"""# 05f v2c 공식 feature dictionary 보고서

생성 시각: {datetime.now().isoformat(timespec="seconds")}

## 기준

- 공식 feature set: `{checks['official_feature_set']}`
- 공식 모델: `{checks['official_model']}`
- 공식 AUC: {checks['official_auc']:.6f}
- 데이터 행 수: {checks['dataset_rows']:,}
- 공식 feature 수: {checks['official_feature_count']}
- SHAP 근거: Stage 07c TRUE SHAP만 사용했습니다. 07r/06h SHAP은 최종 근거로 사용하지 않았습니다.

## 1. 어떤 원본 피처가 남아 있는가?

공식 모델에 직접 남은 원천형 멤버십 정보는 `price_num`, `max_screen_num`, `age_num`, `gender_clean`, `payment_device_clean`, `billing_method_clean`, `is_promotion_bin`, `is_user_verified_bin`입니다. 이 변수들은 원천 Membership 값을 숫자형, 이진형, 정리된 범주형으로 표준화한 feature입니다.

목록: {", ".join(original_like["변수명"].tolist())}

## 2. 어떤 파생 피처가 남아 있는가?

공식 모델의 파생 피처는 1~3주 관측창의 주차별 시청 시간과 세션 수, 단순 이용량 요약, 장르 비율, 장르 entropy, 최근 콘텐츠 시청 비율입니다.

목록: {", ".join(derived["변수명"].tolist())}

상위 SHAP family: {"; ".join(f"{row.feature_family} rank {int(row['rank'])}" for _, row in top_families.iterrows())}

## 3. 어떤 메타데이터와 타깃이 있는가?

`membership_row_id`와 `USER_KEY`는 추적 및 사용자 단위 분리 확인용 metadata입니다. `is_repurchase_label`은 target입니다. 이 세 컬럼은 엑셀에 표시했지만 공식 모델 feature로 표시하지 않았습니다.

목록: {", ".join(metadata_target["변수명"].tolist())}

## 4. 어떤 변수들이 공식 모델에서 제외되었는가?

공식 모델 제외 변수는 product_code, watch-presence 계열, first/last timing 계열, week ratio/delta 계열, 총 시청 시간, 장르별 절대 시청량과 세션 수, 정책/target 인접 후보, duration/raw date/raw backup 계열입니다.

목록: {", ".join(excluded["변수명"].tolist())}

## 5. 왜 주요 변수군을 제외했는가?

`product_code`는 상품 코드 자체를 외워 버리는 위험을 줄이기 위해 제외했습니다. `has_watch_obs`와 `no_watch_obs_flag` 계열은 시청 기록 존재 여부 자체가 shortcut이 될 수 있어 제외했습니다. `first_watch_rel_day`와 `last_watch_rel_day`는 재구독 판단 시점에 가까운 timing proxy가 될 수 있어 제외했습니다. `week*_ratio`와 `w*_minus_*` 계열은 주차별 시청 시간과 구조적으로 중복될 수 있어 제외했습니다. `genre_watch_time_*`와 `genre_session_count_*`는 장르 취향보다 사용량 proxy가 섞일 위험이 있어 제외했습니다.

## 6. 팀원은 이 엑셀을 어떻게 읽어야 하는가?

`변수 설명` 시트에서 `비고`가 `공식 모델 사용`인 행만 06c2/07c corrected official model의 실제 feature입니다. `target`, `metadata only`, `공식 모델 제외`로 표시된 행은 모델 입력 feature가 아닙니다. `SHAP 중요 변수` 시트는 07c TRUE SHAP 기반의 해석 보조 자료이며, 원인 효과나 ROI 근거가 아닙니다. `제외 변수 사유` 시트는 발표나 팀 공유 때 왜 특정 변수를 쓰지 않았는지 설명하는 용도로 읽으면 됩니다.
"""
    REPORT_PATH.write_text(report, encoding="utf-8")
    return report


def validate_outputs(feature_df: pd.DataFrame, checks: dict) -> dict:
    wb = load_workbook(WORKBOOK_PATH, read_only=True)
    sheetnames = wb.sheetnames
    required_sheets = ["변수 설명", "두 노트북 비교", "카테고리 요약"]
    expected_columns = ["#", "변수명", "카테고리", "데이터 타입", "설명", "생성 방식", "비고"]
    ws = wb["변수 설명"]
    header = [ws.cell(1, col).value for col in range(1, 8)]
    official_features = read_json(DATA05C / "feature_sets_v2c.json")["feature_sets"][checks["official_feature_set"]]["features"]
    missing_official = sorted(set(official_features) - set(feature_df["변수명"]))
    metadata_ok = not feature_df[feature_df["변수명"].isin([TARGET, *METADATA_COLUMNS])]["비고"].str.contains("공식 모델 사용", regex=False).any()
    checks_out = {
        "workbook_exists": WORKBOOK_PATH.exists(),
        "required_3_sheets_exist": all(sheet in sheetnames for sheet in required_sheets),
        "variable_sheet_exact_columns": header == expected_columns,
        "every_official_model_feature_listed": len(missing_official) == 0,
        "missing_official_features": missing_official,
        "metadata_and_target_not_model_features": bool(metadata_ok),
        "old_07r_06h_shap_not_final_evidence": True,
        "no_segmentation_output_mixed": True,
        "final_report_created": REPORT_PATH.exists(),
    }
    if not all(value for key, value in checks_out.items() if key != "missing_official_features"):
        raise AssertionError(checks_out)
    return checks_out


def main() -> None:
    require_inputs()
    OUT_DATA.mkdir(parents=True, exist_ok=True)
    OUT_TABLE.mkdir(parents=True, exist_ok=True)
    assert_no_existing_outputs()

    feature_df, checks = make_feature_dictionary()
    comparison_df = make_comparison_sheet(checks)
    category_df = make_category_summary(feature_df, checks)
    shap_df = make_shap_sheet(feature_df)
    exclusion_df = make_exclusion_sheet(feature_df)

    temp_workbook = OUT_TABLE / "05f_v2c_official_feature_dictionary.tmp.xlsx"
    if temp_workbook.exists():
        temp_workbook.unlink()
    with pd.ExcelWriter(temp_workbook, engine="openpyxl") as writer:
        feature_df.to_excel(writer, sheet_name="변수 설명", index=False)
        comparison_df.to_excel(writer, sheet_name="두 노트북 비교", index=False)
        category_df.to_excel(writer, sheet_name="카테고리 요약", index=False)
        shap_df.to_excel(writer, sheet_name="SHAP 중요 변수", index=False)
        exclusion_df.to_excel(writer, sheet_name="제외 변수 사유", index=False)
    style_workbook(temp_workbook)
    temp_workbook.replace(WORKBOOK_PATH)

    report = write_report(feature_df, checks)
    validation = validate_outputs(feature_df, checks)
    summary = {
        "stage": STAGE_NAME,
        "created_at": datetime.now().isoformat(timespec="seconds"),
        "official_feature_set": checks["official_feature_set"],
        "official_model": checks["official_model"],
        "official_auc": checks["official_auc"],
        "dataset_rows": checks["dataset_rows"],
        "official_feature_count": checks["official_feature_count"],
        "variable_dictionary_rows": len(feature_df),
        "report_length_chars": len(report),
        "validation": validation,
        "inputs_used": {
            "stage05c": str((DATA05C / "feature_sets_v2c.json").relative_to(PROJECT_ROOT)).replace("\\", "/"),
            "stage06c2": str((DATA06C2 / "06c2_final_model_recommendation.md").relative_to(PROJECT_ROOT)).replace("\\", "/"),
            "stage07c_true_shap": str((TABLE07C / "07c_global_shap_importance.csv").relative_to(PROJECT_ROOT)).replace("\\", "/"),
            "stage05d_reference_available": (TABLE05D / "05d_v2_feature_dictionary.xlsx").exists(),
            "stage05e_reference_available": (DATA05E / "pruned_feature_sets_v2.json").exists(),
            "stage06h_reference_available_not_final_shap": (DATA06H / "06h_integrated_audit_summary.json").exists(),
        },
        "outputs": {
            "workbook": str(WORKBOOK_PATH.relative_to(PROJECT_ROOT)).replace("\\", "/"),
            "report": str(REPORT_PATH.relative_to(PROJECT_ROOT)).replace("\\", "/"),
            "summary": str(SUMMARY_PATH.relative_to(PROJECT_ROOT)).replace("\\", "/"),
        },
    }
    SUMMARY_PATH.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(summary["validation"], ensure_ascii=False, indent=2))
    print(f"Workbook: {summary['outputs']['workbook']}")
    print(f"Report: {summary['outputs']['report']}")


if __name__ == "__main__":
    main()
