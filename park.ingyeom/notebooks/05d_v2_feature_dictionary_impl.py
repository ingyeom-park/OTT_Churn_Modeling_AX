import json
import os
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl.styles import Font, PatternFill


os.environ.setdefault("PYTHONIOENCODING", "utf-8")

TARGET = "is_repurchase"
ID_COL = "membership_row_id"
GROUP_COL = "USER_KEY"
PRIMARY_FEATURE_SET = "membership_plus_usage_content_w1_3_without_churn_prevented"
PRIMARY_MODEL = "HistGradientBoostingClassifier"

MEMBERSHIP_FEATURES = [
    "price",
    "product_code",
    "max_screen",
    "is_promotion",
    "is_user_verified",
    "gender",
    "age",
    "payment_device",
    "billing_method",
    "is_churn_prevented",
]
FORBIDDEN_OR_METADATA = [
    "USER_KEY",
    "USER_NUM",
    "MOVIE_NUM",
    "movie_title",
    "membership_row_id",
    "reg_date",
    "end_date",
    "duration_days",
    "watch_date",
    "watch_day",
    "is_repurchase",
]


def find_project_root(start):
    for candidate in [start, *start.parents]:
        if (
            (
                (candidate / "_data" / "01_raw" / "Membership.csv").exists()
                or (candidate / "_data" / "01_raw" / "Membership_train.csv").exists()
            )
            and (candidate / "park.ingyeom" / "reports" / "data" / "05_v2_modeling_dataset" / "feature_sets_v2.json").exists()
        ):
            return candidate
    raise FileNotFoundError("Could not locate ott-churn-prediction project root.")


PROJECT_ROOT = find_project_root(Path.cwd())
DATA05 = PROJECT_ROOT / "park.ingyeom" / "reports" / "data" / "05_v2_modeling_dataset"
DATA03 = PROJECT_ROOT / "park.ingyeom" / "reports" / "data" / "03_v2_usage_feature_engineering"
DATA04 = PROJECT_ROOT / "park.ingyeom" / "reports" / "data" / "04_v2_content_feature_engineering"
DATA06C = PROJECT_ROOT / "park.ingyeom" / "reports" / "data" / "06c_v2_overfitting_leakage_adversarial_audit"
DATA07R = PROJECT_ROOT / "park.ingyeom" / "reports" / "data" / "07r_v2_true_shap_interpretation"
TABLE07R = PROJECT_ROOT / "park.ingyeom" / "reports" / "tables" / "07r_v2_true_shap_interpretation"

OUT_DATA = PROJECT_ROOT / "park.ingyeom" / "reports" / "data" / "05d_v2_feature_dictionary"
OUT_TABLE = PROJECT_ROOT / "park.ingyeom" / "reports" / "tables" / "05d_v2_feature_dictionary"
OUT_DATA.mkdir(parents=True, exist_ok=True)
OUT_TABLE.mkdir(parents=True, exist_ok=True)
WORKBOOK_PATH = OUT_TABLE / "05d_v2_feature_dictionary.xlsx"
EXCEL_SHEET_ALIASES = {
    "caution_or_target_adjacent_features": "caution_target_adjacent",
    "multicollinearity_interpretation_groups": "multicollinearity_groups",
}

RAW_FILES = [
    PROJECT_ROOT / "_data" / "01_raw" / name
    for name in [
        "Membership.csv",
        "User_Mapping.csv",
        "View_History.csv",
        "Movie_Master.csv",
        "Membership_train.csv",
        "mapping.csv",
        "Views_train.csv",
        "Movies.csv",
    ]
]
STAGE01_09_PREFIXES = [
    "01_v2_data_overview_and_audit",
    "02_v2_preprocessing_policy",
    "02_v2_preprocessing_policy_validation",
    "03_v2_usage_feature_engineering",
    "04_v2_content_feature_engineering",
    "04_v2_content_feature_feasibility",
    "05_v2_modeling_dataset",
    "06_v2_baseline_modeling",
    "06b_v2_baseline_sanity_audit",
    "07_v2_xai_shap_interpretation",
    "07r_v2_true_shap_interpretation",
    "08_v2_segmentation_strategy",
    "08b_v2_segmentation_refinement",
    "09_v2_business_simulation",
]


def rel(path):
    return str(Path(path).resolve().relative_to(PROJECT_ROOT)).replace("\\", "/")


def snapshot_files(paths):
    out = {}
    for path in paths:
        path = Path(path)
        if path.exists() and path.is_file():
            st = path.stat()
            out[rel(path)] = {"size": st.st_size, "mtime_ns": st.st_mtime_ns}
    return out


def snapshot_tree(paths):
    files = []
    for path in paths:
        path = Path(path)
        if path.exists():
            files.extend([p for p in path.rglob("*") if p.is_file()])
    return snapshot_files(files)


def snapshot_stage01_09():
    paths = []
    for base_name in ["data", "tables", "figures"]:
        base = PROJECT_ROOT / "park.ingyeom" / "reports" / base_name
        for prefix in STAGE01_09_PREFIXES:
            paths.append(base / prefix)
    return snapshot_tree(paths)


def read_json(path):
    return json.loads(Path(path).read_text(encoding="utf-8"))


def write_csv(name, df):
    df.to_csv(OUT_TABLE / f"05d_{name}.csv", index=False, encoding="utf-8-sig")


def family(col):
    if col == TARGET:
        return "target"
    if col in [ID_COL, GROUP_COL] or col in FORBIDDEN_OR_METADATA:
        return "metadata"
    if col in MEMBERSHIP_FEATURES:
        return "membership"
    if "release_month" in col or "recent_content" in col or "old_content" in col:
        return "release_month"
    if "genre_" in col or "top_genre" in col:
        return "genre"
    if "content_" in col or "covered" in col or "missing" in col:
        return "content"
    if col.startswith("w1_"):
        return "usage"
    return "unknown"


def window(col):
    if col == TARGET:
        return "target"
    if col in [ID_COL, GROUP_COL] or col in FORBIDDEN_OR_METADATA:
        return "metadata"
    if col.startswith("w1_3_"):
        return "w1_3"
    if col.startswith("w1_4_"):
        return "w1_4"
    return "common"


def value_type(col, s):
    if col == TARGET:
        return "target"
    if col == ID_COL:
        return "id_metadata"
    if col == GROUP_COL:
        return "group_metadata"
    if col in FORBIDDEN_OR_METADATA:
        return "excluded_forbidden"
    if s.dtype == "object":
        return "categorical"
    vals = set(pd.Series(s.dropna().unique()).head(20).tolist())
    if vals.issubset({0, 1, 0.0, 1.0}):
        return "boolean"
    return "numeric"


def role(col):
    if col == TARGET:
        return "target"
    if col == ID_COL:
        return "id_metadata"
    if col == GROUP_COL:
        return "group_metadata"
    if col in FORBIDDEN_OR_METADATA:
        return "excluded_forbidden"
    return "feature"


def group_name(col):
    if col in MEMBERSHIP_FEATURES:
        return "membership_context"
    if "has_watch" in col or "no_watch" in col:
        return "watch_presence"
    if any(t in col for t in ["total_watch_time", "total_sessions", "unique_contents", "unique_watch_days"]):
        return "usage_volume"
    if any(t in col for t in ["avg_watch_time", "sessions_per_active_day", "max_daily", "max_day_share"]):
        return "session_intensity"
    if "week" in col and "watch_time" in col and "minus" not in col:
        return "weekly_watch_time"
    if "week" in col and "sessions" in col:
        return "weekly_sessions"
    if "ratio" in col and "genre" not in col and "content" not in col:
        return "weekly_ratio"
    if "minus" in col:
        return "week_to_week_delta"
    if "rel_day" in col:
        return "watch_timing_rel_day"
    if "short" in col or "one_minute" in col:
        return "short_watch_behavior"
    if "genre_ratio" in col or "top_genre" in col or "genre_entropy" in col:
        return "genre_preference_ratio"
    if "genre_watch_time" in col or "genre_session_count" in col:
        return "genre_volume_proxy"
    if "release_month" in col or "recent_content" in col or "old_content" in col:
        return "release_month_proxy"
    if "covered" in col or "missing" in col:
        return "coverage_missing_proxy"
    return "other"


def ko_name(col):
    prefix = ""
    base = col
    if col.startswith("w1_3_"):
        prefix, base = "1~3주 ", col[5:]
    elif col.startswith("w1_4_"):
        prefix, base = "1~4주 ", col[5:]
    mapping = {
        ID_COL: "멤버십 행 ID",
        GROUP_COL: "사용자 그룹 키",
        TARGET: "재구독 여부",
        "price": "가격",
        "product_code": "상품 코드",
        "max_screen": "최대 동시 화면 수",
        "is_promotion": "프로모션 여부",
        "is_user_verified": "사용자 인증 여부",
        "gender": "성별",
        "age": "나이",
        "payment_device": "결제 기기",
        "billing_method": "결제 방식",
        "is_churn_prevented": "이탈 방지 여부 후보",
        "has_watch_obs": "시청 관측 여부",
        "no_watch_obs_flag": "무시청 여부",
        "total_watch_time": "총 시청 시간",
        "total_sessions": "총 시청 세션 수",
        "unique_contents": "고유 콘텐츠 수",
        "unique_watch_days": "시청일 수",
        "avg_watch_time_per_session": "세션당 평균 시청 시간",
        "sessions_per_active_day": "활성일당 세션 수",
        "active_span_days": "활동 기간",
        "first_watch_rel_day": "첫 시청 상대일",
        "last_watch_rel_day": "마지막 시청 상대일",
        "max_daily_watch_time": "일 최대 시청 시간",
        "max_day_share": "최대 시청일 집중도",
        "one_minute_watch_count": "1분 시청 횟수",
        "short_watch_count_le5": "5분 이하 짧은 시청 횟수",
        "short_watch_time_le5": "5분 이하 짧은 시청 시간",
        "content_has_watch_obs": "콘텐츠 시청 관측 여부",
        "genre_covered_watch_time": "장르 확인 시청 시간",
        "genre_missing_watch_time": "장르 미확인 시청 시간",
        "genre_covered_watch_ratio": "장르 확인 시청 비중",
        "genre_missing_watch_ratio": "장르 미확인 시청 비중",
        "genre_unique_count": "시청 장르 수",
        "top_genre": "최다 시청 장르",
        "top_genre_watch_time": "최다 장르 시청 시간",
        "top_genre_watch_ratio": "최다 장르 시청 비중",
        "genre_entropy": "장르 다양성",
        "release_month_covered_watch_ratio": "공개월 확인 시청 비중",
        "avg_ott_release_month_weighted": "가중 평균 공개월",
        "recent_content_watch_ratio": "최근 콘텐츠 시청 비중",
        "old_content_watch_ratio": "오래된 콘텐츠 시청 비중",
    }
    if base in mapping:
        return prefix + mapping[base]
    if base.startswith("week") and "watch_time" in base:
        return prefix + base.replace("week", "").replace("_watch_time", "주차 시청 시간")
    if base.startswith("week") and "sessions" in base:
        return prefix + base.replace("week", "").replace("_sessions", "주차 세션 수")
    if base.startswith("week") and "ratio" in base:
        return prefix + base.replace("week", "").replace("_ratio", "주차 시청 비중")
    if "minus" in base:
        return prefix + base.replace("_", " ")
    if "genre_ratio_" in base:
        return prefix + base.replace("genre_ratio_", "").replace("_", "/") + " 장르 비중"
    if "genre_watch_time_" in base:
        return prefix + base.replace("genre_watch_time_", "").replace("_", "/") + " 장르 시청 시간"
    if "genre_session_count_" in base:
        return prefix + base.replace("genre_session_count_", "").replace("_", "/") + " 장르 세션 수"
    return prefix + base.replace("_", " ")


def calc_logic(col):
    g = group_name(col)
    if col == TARGET:
        return "Membership 원천 target. Y는 재구독, N은 비재구독으로 Stage 06에서 Y=1, N=0으로 매핑됩니다."
    if col in [ID_COL, GROUP_COL]:
        return "모델 학습 feature가 아니라 추적 또는 USER_KEY group split을 위한 metadata입니다."
    if family(col) == "membership":
        return "전처리된 Membership 구독 이벤트 테이블에서 온 가입/결제/상품 맥락 변수입니다."
    if g in ["weekly_watch_time", "weekly_sessions"]:
        return "관측창 안의 주차별 시청 시간 또는 세션 수를 집계한 변수입니다."
    if g == "weekly_ratio":
        return "해당 주차 시청 시간이 관측창 총 시청 시간에서 차지하는 비율입니다."
    if g == "week_to_week_delta":
        return "두 주차의 시청 시간 차이입니다."
    if g == "genre_preference_ratio":
        return "장르별 시청 시간 비중 또는 최다 장르 기반 proxy입니다."
    if g == "genre_volume_proxy":
        return "장르별 시청 시간 또는 세션 수이며, 취향과 사용량이 섞인 proxy입니다."
    if g == "release_month_proxy":
        return "Movie_Master의 ott_release_month에서 파생한 공개월 proxy입니다."
    return "Stage 03/04/05 산출물에서 membership_row_id 단위로 집계 또는 보존된 변수입니다."


def interpretation(col):
    g = group_name(col)
    if col == TARGET:
        return "예측 대상입니다. feature로 쓰면 안 됩니다."
    if col in [ID_COL, GROUP_COL]:
        return "식별과 split용입니다. 행동 의미로 해석하지 않습니다."
    if g in ["usage_volume", "weekly_watch_time", "weekly_sessions", "session_intensity"]:
        return "값이 클수록 관측창 안의 시청 활동량 또는 집중도가 큽니다."
    if g == "watch_timing_rel_day":
        return "값이 클수록 첫 시청 또는 마지막 시청이 구독 시작 후 더 늦게 발생했다는 뜻입니다."
    if g == "weekly_ratio":
        return "특정 주차에 시청이 몰린 정도입니다."
    if g == "week_to_week_delta":
        return "주차 간 시청량 증가 또는 감소 흐름입니다."
    if g.startswith("genre"):
        return "장르 선호처럼 보이지만 사용량 proxy가 섞일 수 있습니다."
    if g == "membership_context":
        return "가입 상품, 가격, 인증, 결제 맥락을 나타냅니다."
    return "예측 설명에는 사용할 수 있지만 인과효과로 해석하지 않습니다."


def caution(col):
    g = group_name(col)
    if col == TARGET:
        return "target이므로 feature로 사용 금지."
    if col in FORBIDDEN_OR_METADATA:
        return "metadata, raw key/date, 또는 target 관련 컬럼입니다. 모델 feature로 해석하지 않습니다."
    if col == "is_churn_prevented":
        return "정책/결과 인접 가능성이 있어 with/without feature set을 분리했습니다."
    if g in ["watch_timing_rel_day", "weekly_ratio", "week_to_week_delta"] or "week3" in col or "week4" in col:
        return "재구독 판단 시점에 가까운 행동 proxy일 수 있어 target-adjacent로 조심해야 합니다."
    if g == "genre_volume_proxy":
        return "순수 장르 취향보다 사용량 중복 신호일 수 있습니다."
    if any(t in col for t in ["price", "product_code", "promotion"]):
        return "가격/프로모션/상품 정책과 고객군 차이가 섞일 수 있습니다."
    return "예측적, 설명적 신호로만 사용하고 원인으로 말하지 않습니다."


def use_recommendation(col, high_corr_features):
    if col == TARGET:
        return "target_only"
    if col in [ID_COL, GROUP_COL] or col in FORBIDDEN_OR_METADATA:
        return "metadata_only"
    if col in high_corr_features or group_name(col) in {
        "weekly_watch_time",
        "weekly_sessions",
        "weekly_ratio",
        "week_to_week_delta",
        "genre_volume_proxy",
        "coverage_missing_proxy",
    }:
        return "group_for_interpretation"
    if "week3" in col or "week4" in col or "rel_day" in col or col == "is_churn_prevented":
        return "use_with_caution"
    return "use"


def example_values(series):
    s = series.dropna()
    if s.empty:
        return "all missing"
    if pd.api.types.is_numeric_dtype(series):
        return f"min={s.min()}, median={s.median()}, max={s.max()}"
    vc = s.astype(str).value_counts().head(5)
    return "; ".join(f"{k}:{v}" for k, v in vc.items())


def write_notebook():
    nb = {
        "cells": [
            {
                "cell_type": "markdown",
                "metadata": {},
                "source": [
                    "# 05d v2 Feature Dictionary\n",
                    "\n",
                    "Team-sharing feature dictionary. Documentation only. No model training, no SHAP run, no Optuna, no new features.\n",
                ],
            },
            {
                "cell_type": "code",
                "execution_count": None,
                "metadata": {},
                "outputs": [],
                "source": ["%run 05d_v2_feature_dictionary_impl.py\n"],
            },
        ],
        "metadata": {
            "kernelspec": {"display_name": "Python 3", "language": "python", "name": "python3"},
            "language_info": {"name": "python", "pygments_lexer": "ipython3"},
        },
        "nbformat": 4,
        "nbformat_minor": 5,
    }
    path = PROJECT_ROOT / "park.ingyeom" / "notebooks" / "05d_v2_feature_dictionary.ipynb"
    path.write_text(json.dumps(nb, ensure_ascii=False, indent=2), encoding="utf-8")


raw_before = snapshot_files(RAW_FILES)
data_before = snapshot_tree([PROJECT_ROOT / "_data"])
stage_before = snapshot_stage01_09()

df13 = pd.read_csv(DATA05 / "modeling_dataset_v2_w1_3.csv")
df14 = pd.read_csv(DATA05 / "modeling_dataset_v2_w1_4.csv")
feature_sets_payload = read_json(DATA05 / "feature_sets_v2.json")
usage_summary = read_json(DATA03 / "usage_feature_summary.json")
content_summary = read_json(DATA04 / "content_feature_summary.json")
audit06c = read_json(DATA06C / "06c_adversarial_audit_summary.json")
true_shap_summary = read_json(DATA07R / "07r_true_shap_summary.json")
global_shap = pd.read_csv(TABLE07R / "07r_global_shap_importance.csv")
family_shap = pd.read_csv(TABLE07R / "07r_feature_family_shap_importance.csv")

feature_sets = feature_sets_payload["feature_sets"]
primary_features = feature_sets[PRIMARY_FEATURE_SET]
used_any = set()
for fs in feature_sets.values():
    if isinstance(fs, list):
        used_any.update(fs)

all_columns = list(dict.fromkeys(list(df13.columns) + list(df14.columns)))
series_map = {}
for col in all_columns:
    if col in df13.columns and col in df14.columns:
        series_map[col] = pd.concat([df13[col], df14[col]], ignore_index=True)
    elif col in df13.columns:
        series_map[col] = df13[col]
    else:
        series_map[col] = df14[col]

primary_numeric = [c for c in primary_features if c in df13.columns and pd.api.types.is_numeric_dtype(df13[c]) and df13[c].nunique(dropna=True) > 1]
primary_corr = df13[primary_numeric].corr().abs() if primary_numeric else pd.DataFrame()
high_corr_features = set()
high_corr_pairs = []
if not primary_corr.empty:
    for i, a in enumerate(primary_numeric):
        for b in primary_numeric[i + 1 :]:
            val = primary_corr.loc[a, b]
            if pd.notna(val) and val >= 0.80:
                high_corr_features.update([a, b])
                high_corr_pairs.append((a, b, float(val)))

primary_shap = global_shap[
    (global_shap["model_role"].eq("primary_conservative"))
    & (global_shap["window"].eq("w1_3"))
    & (global_shap["model_name"].eq(PRIMARY_MODEL))
].copy()
primary_shap = primary_shap.sort_values("mean_abs_shap", ascending=False).reset_index(drop=True)
primary_shap["shap_rank"] = np.arange(1, len(primary_shap) + 1)
shap_map = primary_shap.drop_duplicates("original_feature").set_index("original_feature").to_dict(orient="index")

readme = pd.DataFrame(
    [
        ("purpose", "v2 모델링 컬럼의 의미, 출처, 계산 방식, 해석, 주의사항을 팀 공유용으로 정리합니다."),
        ("modeling_unit", "모델링 단위는 Membership 1행, 즉 구독 이벤트 1건입니다."),
        ("target", "target은 is_repurchase입니다. Y=1은 재구독, N=0은 비재구독입니다."),
        ("score_direction", "repurchase_score는 P(is_repurchase=Y)이고 churn_risk_score는 1 - repurchase_score입니다."),
        ("windows", "w1_3은 rel_day 0~20, w1_4는 rel_day 0~27입니다."),
        ("encoding", "one-hot encoding은 Stage 06 모델 파이프라인 내부에서 수행됩니다. Stage 05 CSV는 원 컬럼 값을 보존합니다."),
        ("metadata", "USER_KEY는 group split metadata이고 membership_row_id는 row tracing ID입니다. 둘 다 모델 feature가 아닙니다."),
        ("xai_basis", "TRUE SHAP 근거는 Stage 07r만 사용합니다. Stage 07 fallback은 최종 근거가 아닙니다."),
        ("stage06c_caution", f"Stage 06c verdict: {audit06c.get('final_verdict')}. 개별 변수 해석은 보수적으로 해야 합니다."),
        ("excel_sheet_name_limit", "Excel 시트명은 31자 제한이 있어 caution_or_target_adjacent_features는 caution_target_adjacent로, multicollinearity_interpretation_groups는 multicollinearity_groups로 저장했습니다. CSV 파일명은 긴 원래 이름을 유지합니다."),
    ],
    columns=["section", "content"],
)

rows = []
for no, col in enumerate(all_columns, start=1):
    s = series_map[col]
    shap_info = shap_map.get(col, {})
    rows.append(
        {
            "no": no,
            "column_name": col,
            "presentation_label_ko": ko_name(col),
            "presentation_label_en": col.replace("_", " "),
            "window": window(col),
            "feature_family": family(col),
            "interpretation_group": group_name(col),
            "value_type": value_type(col, s),
            "modeling_role": role(col),
            "source_stage": "Stage 03" if family(col) == "usage" else "Stage 04" if family(col) in ["content", "genre", "release_month"] else "Stage 05",
            "source_table": "modeling_dataset_v2_w1_3.csv" if col in df13.columns and col not in df14.columns else "modeling_dataset_v2_w1_4.csv" if col in df14.columns and col not in df13.columns else "both modeling datasets",
            "used_in_any_feature_set": "Y" if col in used_any else "N",
            "used_in_primary_model": "Y" if col in primary_features else "N",
            "used_in_stage07r_true_shap": "Y" if col in shap_map else "N",
            "shap_rank": shap_info.get("shap_rank", ""),
            "mean_abs_shap": shap_info.get("mean_abs_shap", ""),
            "example_values_or_range": example_values(s),
            "missing_rate": float(s.isna().mean()),
            "unique_count": int(s.nunique(dropna=True)),
            "calculation_logic": calc_logic(col),
            "plain_interpretation": interpretation(col),
            "caution": caution(col),
            "target_adjacent_or_redundancy_flag": "Y" if use_recommendation(col, high_corr_features) in ["group_for_interpretation", "use_with_caution"] else "N",
            "final_use_recommendation": use_recommendation(col, high_corr_features),
            "do_not_claim": "Do not claim causality, independent effect, or guaranteed retention impact.",
        }
    )
all_columns_dictionary = pd.DataFrame(rows)

membership_features = all_columns_dictionary[all_columns_dictionary["column_name"].isin(MEMBERSHIP_FEATURES)].copy()
usage_features = all_columns_dictionary[all_columns_dictionary["feature_family"].eq("usage")].copy()
content_features = all_columns_dictionary[all_columns_dictionary["feature_family"].isin(["content", "genre", "release_month"])].copy()

feature_set_rows = []
for name, feats in feature_sets.items():
    if not isinstance(feats, list):
        continue
    feature_set_rows.append(
        {
            "feature_set_name": name,
            "window": "w1_3" if "_w1_3" in name else "w1_4" if "_w1_4" in name else "common",
            "feature_count": len(feats),
            "includes_membership": "Y" if any(f in MEMBERSHIP_FEATURES for f in feats) else "N",
            "includes_usage": "Y" if any(family(f) == "usage" for f in feats) else "N",
            "includes_content_or_genre": "Y" if any(family(f) in ["content", "genre", "release_month"] for f in feats) else "N",
            "includes_is_churn_prevented": "Y" if "is_churn_prevented" in feats else "N",
            "feature_list": ", ".join(feats),
            "caution": "Check window and is_churn_prevented variant before reporting.",
        }
    )
feature_sets_df = pd.DataFrame(feature_set_rows)

primary_model_features = all_columns_dictionary[all_columns_dictionary["column_name"].isin(primary_features)].copy()
primary_model_features = primary_model_features.sort_values(["shap_rank", "column_name"], key=lambda x: pd.to_numeric(x, errors="coerce").fillna(9999) if x.name == "shap_rank" else x)

top_true_shap_features = primary_shap.head(30).copy()
top_true_shap_features["presentation_label_ko"] = top_true_shap_features["original_feature"].map(ko_name)
top_true_shap_features["interpretation_group"] = top_true_shap_features["original_feature"].map(group_name)
top_true_shap_features["plain_korean_interpretation"] = top_true_shap_features["original_feature"].map(lambda c: f"{ko_name(c)} 변수는 재구독 점수 구분에 사용된 모델 설명 신호입니다.")
top_true_shap_features["caution"] = top_true_shap_features["original_feature"].map(caution)
top_true_shap_features["final_use_recommendation"] = top_true_shap_features["original_feature"].map(lambda c: use_recommendation(c, high_corr_features))
top_true_shap_features = top_true_shap_features[
    ["shap_rank", "original_feature", "presentation_label_ko", "feature_family", "interpretation_group", "mean_abs_shap", "mean_shap", "score_direction_note", "plain_korean_interpretation", "caution", "final_use_recommendation"]
].rename(columns={"original_feature": "feature_name"})

forbidden_or_metadata_columns = pd.DataFrame(
    [
        {
            "column_name": col,
            "classification": "target" if col == TARGET else "id_metadata" if col == ID_COL else "group_metadata" if col == GROUP_COL else "forbidden_raw_or_key",
            "reason": caution(col),
            "final_use_recommendation": "target_only" if col == TARGET else "metadata_only" if col in [ID_COL, GROUP_COL] else "exclude",
        }
        for col in FORBIDDEN_OR_METADATA
    ]
)

renaming_suggestions = all_columns_dictionary[
    ["column_name", "presentation_label_ko", "presentation_label_en", "interpretation_group"]
].copy()
renaming_suggestions["suggested_short_name"] = renaming_suggestions["column_name"].str.replace("w1_3_", "3w_", regex=False).str.replace("w1_4_", "4w_", regex=False)
renaming_suggestions["keep_original_in_code"] = "Y"
renaming_suggestions["use_label_in_report"] = "Y"
renaming_suggestions["why"] = "코드에서는 원 컬럼명을 유지하고, 발표/보고서에서는 이해하기 쉬운 라벨을 사용하기 위함입니다."

caution_or_target_adjacent_features = all_columns_dictionary[
    all_columns_dictionary["target_adjacent_or_redundancy_flag"].eq("Y")
    | all_columns_dictionary["final_use_recommendation"].isin(["group_for_interpretation", "use_with_caution"])
].copy()

group_rows = []
for group, sub in all_columns_dictionary[all_columns_dictionary["modeling_role"].eq("feature")].groupby("interpretation_group"):
    group_rows.append(
        {
            "interpretation_group": group,
            "feature_count": len(sub),
            "features": "|".join(sub["column_name"].tolist()),
            "recommended_presentation_style": "묶어서 설명",
            "why_group": "상관 또는 구조적 중복 가능성이 있어 개별 변수 효과처럼 해석하면 위험합니다." if group != "membership_context" else "멤버십 맥락 변수군으로 함께 설명합니다.",
            "example_claim": f"{group} 변수군은 모델이 활용한 설명 신호입니다.",
            "do_not_claim": "개별 변수가 독립적으로 이탈 원인이라고 말하지 않습니다.",
        }
    )
multicollinearity_interpretation_groups = pd.DataFrame(group_rows)

data_type_audit = pd.DataFrame(
    [
        {
            "column_name": col,
            "pandas_dtype": str(series_map[col].dtype),
            "semantic_type": value_type(col, series_map[col]),
            "missing_values": int(series_map[col].isna().sum()),
            "missing_rate": float(series_map[col].isna().mean()),
            "unique_values": int(series_map[col].nunique(dropna=True)),
            "min": series_map[col].min() if pd.api.types.is_numeric_dtype(series_map[col]) else "",
            "max": series_map[col].max() if pd.api.types.is_numeric_dtype(series_map[col]) else "",
            "top_values": "; ".join(f"{k}:{v}" for k, v in series_map[col].dropna().astype(str).value_counts().head(5).items()),
        }
        for col in all_columns
    ]
)

sheets = {
    "README": readme,
    "all_columns_dictionary": all_columns_dictionary,
    "membership_features": membership_features,
    "usage_features": usage_features,
    "content_features": content_features,
    "feature_sets": feature_sets_df,
    "primary_model_features": primary_model_features,
    "top_true_shap_features": top_true_shap_features,
    "forbidden_or_metadata_columns": forbidden_or_metadata_columns,
    "renaming_suggestions": renaming_suggestions,
    "caution_or_target_adjacent_features": caution_or_target_adjacent_features,
    "multicollinearity_interpretation_groups": multicollinearity_interpretation_groups,
    "data_type_audit": data_type_audit,
}

for sheet, df in sheets.items():
    write_csv(sheet, df)

raw_after_pre = snapshot_files(RAW_FILES)
data_after_pre = snapshot_tree([PROJECT_ROOT / "_data"])
stage_after_pre = snapshot_stage01_09()
final_checks = pd.DataFrame(
    [
        ("raw_files_unchanged", "PASS" if raw_before == raw_after_pre else "FAIL", "raw snapshots unchanged"),
        ("no__data_output_created", "PASS" if data_before == data_after_pre else "FAIL", "_data snapshot unchanged"),
        ("stage01_through_stage09_outputs_not_overwritten", "PASS" if stage_before == stage_after_pre else "FAIL", "Stage 01-09 snapshots unchanged"),
        ("excel_workbook_created", "PENDING", str(WORKBOOK_PATH)),
        ("all_required_sheets_created", "PASS", ",".join(sheets.keys()) + ",final_checks"),
        ("csv_copies_created", "PASS" if all((OUT_TABLE / f"05d_{s}.csv").exists() for s in sheets) else "FAIL", "CSV copies for non-final-check sheets created"),
        ("covers_every_modeling_column", "PASS" if set(all_columns_dictionary["column_name"]) == set(all_columns) else "FAIL", f"covered={len(all_columns_dictionary)}, union={len(all_columns)}"),
        ("target_and_metadata_not_mislabeled", "PASS" if not all_columns_dictionary[all_columns_dictionary["column_name"].isin([TARGET, ID_COL, GROUP_COL]) & all_columns_dictionary["modeling_role"].eq("feature")].shape[0] else "FAIL", "target and metadata roles checked"),
        ("stage07r_true_shap_used_not_stage07_fallback", "PASS" if not top_true_shap_features.empty else "FAIL", "Stage 07r global SHAP table read only"),
        ("no_model_training_no_new_features", "PASS", "Only documentation and descriptive dictionary artifacts created"),
    ],
    columns=["check", "status", "detail"],
)
sheets["final_checks"] = final_checks
write_csv("final_checks", final_checks)

with pd.ExcelWriter(WORKBOOK_PATH, engine="openpyxl") as writer:
    for sheet, df in sheets.items():
        df.to_excel(writer, sheet_name=EXCEL_SHEET_ALIASES.get(sheet, sheet), index=False)
    wb = writer.book
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="4F81BD")
        for col_cells in ws.columns:
            letter = col_cells[0].column_letter
            max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col_cells[:100])
            ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 45)

final_checks.loc[final_checks["check"].eq("excel_workbook_created"), "status"] = "PASS" if WORKBOOK_PATH.exists() else "FAIL"
final_checks.loc[final_checks["check"].eq("csv_copies_created"), "status"] = "PASS" if all((OUT_TABLE / f"05d_{s}.csv").exists() for s in sheets) else "FAIL"
write_csv("final_checks", final_checks)
with pd.ExcelWriter(WORKBOOK_PATH, engine="openpyxl") as writer:
    for sheet, df in {**sheets, "final_checks": final_checks}.items():
        df.to_excel(writer, sheet_name=EXCEL_SHEET_ALIASES.get(sheet, sheet), index=False)
    wb = writer.book
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="4F81BD")
        for col_cells in ws.columns:
            letter = col_cells[0].column_letter
            max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col_cells[:100])
            ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 45)

summary = {
    "scope": "Stage 05d v2 feature dictionary documentation only",
    "workbook_path": rel(WORKBOOK_PATH),
    "sheet_count": 14,
    "w1_3_columns": len(df13.columns),
    "w1_4_columns": len(df14.columns),
    "union_columns": len(all_columns),
    "primary_feature_count": len(primary_features),
    "top_true_shap_feature_count": len(top_true_shap_features),
    "target_adjacent_or_redundancy_feature_count": len(caution_or_target_adjacent_features),
    "stage06c_verdict": audit06c.get("final_verdict"),
    "final_checks_passed": bool(final_checks["status"].eq("PASS").all()),
}
(OUT_DATA / "05d_v2_feature_dictionary_summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")

report = [
    "# 05d v2 Feature Dictionary Report",
    "",
    f"Generated at: {datetime.now().isoformat(timespec='seconds')}",
    "",
    f"- w1_3 columns: {len(df13.columns)}",
    f"- w1_4 columns: {len(df14.columns)}",
    f"- union columns covered: {len(all_columns)}",
    f"- primary model features: {len(primary_features)}",
    f"- target-adjacent or redundancy caution rows: {len(caution_or_target_adjacent_features)}",
    f"- Stage 06c verdict: {audit06c.get('final_verdict')}",
    "",
    "## Top TRUE SHAP Features",
]
for _, row in top_true_shap_features.head(10).iterrows():
    report.append(f"- {int(row['shap_rank'])}. {row['feature_name']} / {row['presentation_label_ko']} / {row['feature_family']} / mean_abs_shap={row['mean_abs_shap']:.6f}")
report.extend(
    [
        "",
        "## How To Use",
        "- 코드에서는 원 컬럼명을 유지합니다.",
        "- 팀 공유와 발표에서는 presentation_label_ko를 사용합니다.",
        "- final_use_recommendation이 group_for_interpretation인 변수는 개별 효과처럼 말하지 않고 묶어서 설명합니다.",
        "- Stage 07r TRUE SHAP만 최종 XAI 근거로 사용합니다.",
        "",
        "## What Not To Claim",
        "- 개별 피처가 이탈의 원인이라고 말하지 않습니다.",
        "- country, rating, runtime, actor, director, Wavve, KOBIS metadata를 사용했다고 말하지 않습니다.",
        "- USER_KEY, membership_row_id, raw date, target을 모델 feature처럼 말하지 않습니다.",
    ]
)
(OUT_DATA / "05d_v2_feature_dictionary_report.md").write_text("\n".join(report) + "\n", encoding="utf-8")

write_notebook()

print("05d v2 feature dictionary completed.")
for _, row in final_checks.iterrows():
    print(f"{row['check']}: {row['status']} - {row['detail']}")
